# pip install pyserial openpyxl

import serial
from openpyxl import Workbook, load_workbook
import os

# Serial port (change this to your Arduino port)
PORT = "/dev/ttyUSB0"  # For Linux
# Example for Windows: PORT = "COM3"

BAUDRATE = 9600
FILENAME = "rfid_log.xlsx"

# Check if Excel file exists, else create one
if not os.path.exists(FILENAME):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Time"])  # headers
    wb.save(FILENAME)

# Open serial connection
ser = serial.Serial(PORT, BAUDRATE)
print("Listening on", PORT)

# Load workbook
wb = load_workbook(FILENAME)
ws = wb.active

try:
    while True:
        line = ser.readline().decode('utf-8').strip()
        if line:
            from datetime import datetime
            ws.append([line, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            wb.save(FILENAME)
            print(f"Saved: {line}")
except KeyboardInterrupt:
    print("Stopped by user.")
finally:
    ser.close()
